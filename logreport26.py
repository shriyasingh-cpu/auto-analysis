"""
=============================================================================
  Flight Log Report Generator  —  v3.6
  -----------------------------------------------
  CHANGES FROM v3.5:
    • All takeoff / landing times displayed in IST (UTC +5:30).
      GPS anchors are still computed in UTC; +5h 30m is added only
      at the final formatting step inside timeus_to_datetime_str().

  CHANGES FROM v3.4:
    • FIXED: Takeoff / Landing datetime now derived from GPS time, NOT
      from pymavlink's msg._timestamp (which is unreliable / synthetic).

  HOW DATETIME IS NOW COMPUTED (GPS anchor method)
  ─────────────────────────────────────────────────
  ArduPilot .bin logs do NOT embed real wall-clock time per message.
  The only reliable real-time source is the GPS message, which carries:
      GWk  — GPS week number  (integer)
      GMS  — GPS milliseconds into that week  (integer ms)
      TimeUS — log microseconds at that GPS fix

  STEP 1: After bin_to_csv(), scan GPS.csv for rows where Status >= 3
          (3D fix).  Build a list of anchors:
              anchor = (TimeUS, wall_clock_datetime)
          where wall_clock_datetime is computed from GWk + GMS using the
          standard GPS epoch (1980-01-06 00:00:00 UTC) with leap-second
          correction (18 s as of 2017, good until next IERS bulletin).

  STEP 2: For any target TimeUS (takeoff or landing row), find the
          nearest anchor and compute:
              offset_us = target_TimeUS − anchor_TimeUS
              wall_clock = anchor_datetime + timedelta(microseconds=offset_us)

  STEP 3: Format as  'YYYY-MM-DD HH:MM:SS.mmm'  — exactly what the
          Excel sheet expects.

  This gives accurate wall-clock time to millisecond precision without
  relying on pymavlink's unreliable _timestamp field.

  FULL PIPELINE:
    Step 1 — Browse one or more .bin ArduPilot log files
    Step 2 — Convert each .bin to per-type CSVs via pymavlink
    Step 3 — Extract flight mode, takeoff/landing (ThO), endurance,
              battery, vibration (500ms on-air avg per axis), altitude
    Step 4 — Build colour-coded Excel workbook (one row per BIN)

  Requirements:
    pip install numpy pandas pymavlink openpyxl
=============================================================================
"""

import os, sys, warnings
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict

warnings.filterwarnings("ignore")

try:
    import numpy as np
    NUMPY_OK = True
except ImportError:
    NUMPY_OK = False

try:
    import pandas as pd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False
    print("[ERROR] pandas not installed — pip install pandas")

try:
    from pymavlink import mavutil
    PYMAV_OK = True
except ImportError:
    PYMAV_OK = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("[ERROR] openpyxl not installed — pip install openpyxl")

try:
    import tkinter as tk
    from tkinter import filedialog
    TK_OK = True
except ImportError:
    TK_OK = False


# ═══════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════

THO_FLYING_THRESHOLD  = 0.0
VIBE_SAMPLE_INTERVAL  = 500_000   # 500 ms in µs

VIBE_GOOD    = 10.0
VIBE_OK      = 15.0
VIBE_CAUTION = 20.0
VIBE_WARNING = 30.0

BATT_VOLT_CRIT = 10.5
BATT_VOLT_WARN = 11.1
BATT_CURR_HIGH = 60.0
VOLT_DROP_THR  = 0.5

SKIP_MSG_TYPES = {"BAD_DATA", "FMT", "FMTU", "UNIT", "MULT", "PARM"}

# GPS epoch: 1980-01-06 00:00:00 UTC
GPS_EPOCH = datetime(1980, 1, 6, 0, 0, 0)
# Leap seconds between GPS time and UTC as of 2017 (valid until next IERS bulletin)
GPS_LEAP_SECONDS = 18
# Local timezone offset applied to all displayed takeoff / landing times
# IST = UTC + 5 h 30 m
IST_OFFSET = timedelta(hours=5, minutes=30)

# colours
C_HDR_DARK  = "1A5276"; C_HDR_MID   = "2471A3"; C_HDR_GREEN = "1E8449"
C_WHITE     = "FFFFFF";  C_ALT       = "EBF5FB"
C_GOOD      = "C6EFCE";  C_OK        = "E2EFDA"
C_CAUTION   = "FFEB9C";  C_WARNING   = "FFC7CE"; C_CRITICAL  = "FF4444"
C_VOLT_WARN = "FFE699";  C_VOLT_CRIT = "FF4444"; C_CURR_HIGH = "F4CCCC"
C_TAKEOFF   = "D6EAF8";  C_LANDING   = "FEF9E7"
C_TAKEOFF_H = "1E5F74";  C_LANDING_H = "9E6B21"
C_ENDUR     = "E8F8E8";  C_ENDUR_H   = "1A6B1A"; C_PARTIAL   = "FFD580"

FNT_HDR    = Font(name="Arial", bold=True,  color=C_WHITE,  size=10)
FNT_SUBHDR = Font(name="Arial", bold=True,  color=C_WHITE,  size=9)
FNT_BODY   = Font(name="Arial", bold=False, color="000000", size=9)
FNT_BOLD   = Font(name="Arial", bold=True,  color="000000", size=9)
FNT_RED    = Font(name="Arial", bold=True,  color="C00000", size=9)
FNT_ORANGE = Font(name="Arial", bold=True,  color="974706", size=9)
FNT_YELLOW = Font(name="Arial", bold=True,  color="7D6608", size=9)
FNT_GREEN  = Font(name="Arial", bold=True,  color="1E8449", size=9)
FNT_WHITE  = Font(name="Arial", bold=True,  color=C_WHITE,  size=9)

AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def fill(c):
    if isinstance(c, int): c = f"{c:06X}"
    c = str(c).strip().lstrip("#")
    if len(c) == 6: c = "FF" + c
    from openpyxl.styles.colors import Color
    return PatternFill("solid", fgColor=Color(rgb=c))

def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def thick():
    s = Side(style="medium", color="555555")
    return Border(left=s, right=s, top=s, bottom=s)


# ═══════════════════════════════════════════════════════════════════════════
#  BIN → CSV
# ═══════════════════════════════════════════════════════════════════════════

def bin_to_csv(bin_path, out_dir, log=print):
    """Convert .bin to per-type CSV files. No datetime injection here —
    real wall-clock time is reconstructed later from GPS.csv anchors."""
    if not PYMAV_OK:
        log("[ERROR] pymavlink not installed."); return None, {}
    bin_path = Path(bin_path)
    log(f"  Opening: {bin_path.name}")
    try:
        mlog = mavutil.mavlink_connection(str(bin_path), dialect="ardupilotmega")
    except Exception as e:
        log(f"  [ERROR] {e}"); return None, {}
    data = defaultdict(list); total = 0
    log("  Reading messages…")
    while True:
        try:
            msg = mlog.recv_match(blocking=False)
            if msg is None: break
            mt = msg.get_type()
            if mt in SKIP_MSG_TYPES: continue
            row = msg.to_dict(); row.pop("mavpackettype", None)
            data[mt].append(row); total += 1
            if total % 10_000 == 0:
                log(f"    Read {total:,}  types: {len(data)}")
        except Exception: continue
    log(f"  Done — {total:,} messages  {len(data)} types")
    csv_dir = Path(out_dir) / "csv"; csv_dir.mkdir(parents=True, exist_ok=True)
    csv_files = {}
    for mt, rows in sorted(data.items()):
        if not rows: continue
        try:
            df = pd.DataFrame(rows)
            path = csv_dir / f"{mt}.csv"
            df.to_csv(path, index=False); csv_files[mt] = str(path)
        except Exception as e:
            log(f"  [WARN] {mt}: {e}")
    log(f"  Saved {len(csv_files)} CSV files → {csv_dir}")
    return csv_dir, csv_files


# ═══════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def find_col(df, *names):
    for n in names:
        if n in df.columns: return n
    lc = {c.lower(): c for c in df.columns}
    for n in names:
        if n.lower() in lc: return lc[n.lower()]
    return None

def load(csv_dir, *names):
    csv_dir = Path(csv_dir)
    for n in names:
        p = csv_dir / f"{n}.csv"
        if p.exists():
            try: return pd.read_csv(p)
            except Exception: pass
    return None

def sec_to_dur(secs):
    secs = max(0, int(secs))
    return f"{secs // 60}m {secs % 60:02d}s"


# ═══════════════════════════════════════════════════════════════════════════
#  GPS-BASED DATETIME ANCHOR BUILDER
# ═══════════════════════════════════════════════════════════════════════════

def gps_week_ms_to_utc(gps_week, gps_ms):
    """
    Convert GPS week + milliseconds-into-week → UTC datetime.

    GPS time started at 1980-01-06 00:00:00 UTC.
    GPS time does NOT include leap seconds; UTC does.
    As of 2017 there are 18 leap seconds between GPS and UTC.

    Formula:
        gps_seconds_total = gps_week * 604800 + gps_ms / 1000
        utc = GPS_EPOCH + timedelta(seconds=gps_seconds_total) - leap_seconds
    """
    gps_total_seconds = int(gps_week) * 604800 + int(gps_ms) / 1000.0
    utc = GPS_EPOCH + timedelta(seconds=gps_total_seconds) - timedelta(seconds=GPS_LEAP_SECONDS)
    return utc


def build_gps_time_anchors(csv_dir, log=print):
    """
    Read GPS.csv and build a sorted list of (TimeUS, utc_datetime) anchors.

    Only rows with a valid 3D GPS fix (Status >= 3) and non-zero GWk are used.
    These anchors let us convert any log TimeUS to a real wall-clock time.

    Returns: sorted list of (timeus_int, datetime) tuples, or empty list.
    """
    df_gps = load(csv_dir, "GPS", "GPS2")
    if df_gps is None:
        log("    [WARN] GPS.csv not found — datetime will fall back to TimeUS offset")
        return []

    tus_col  = find_col(df_gps, "TimeUS", "timeus", "Time_us", "Time")
    gwk_col  = find_col(df_gps, "GWk", "GPSWeek", "Week", "week")
    gms_col  = find_col(df_gps, "GMS", "GPSMs",   "TimeMS", "ms")
    stat_col = find_col(df_gps, "Status", "status", "Fix", "fix")

    if tus_col is None or gwk_col is None or gms_col is None:
        log(f"    [WARN] GPS columns not found (tus={tus_col} gwk={gwk_col} gms={gms_col})")
        log(f"    [WARN] GPS.csv columns: {list(df_gps.columns)}")
        return []

    df_gps = df_gps.copy()
    df_gps[tus_col]  = pd.to_numeric(df_gps[tus_col],  errors="coerce")
    df_gps[gwk_col]  = pd.to_numeric(df_gps[gwk_col],  errors="coerce")
    df_gps[gms_col]  = pd.to_numeric(df_gps[gms_col],  errors="coerce")
    if stat_col:
        df_gps[stat_col] = pd.to_numeric(df_gps[stat_col], errors="coerce")

    # Keep only rows with a valid 3D fix and non-zero GPS week
    mask = (
        df_gps[tus_col].notna() &
        df_gps[gwk_col].notna() &
        df_gps[gms_col].notna() &
        (df_gps[gwk_col] > 0)
    )
    if stat_col:
        mask &= (df_gps[stat_col] >= 3)

    df_valid = df_gps[mask].sort_values(tus_col).reset_index(drop=True)

    if df_valid.empty:
        log("    [WARN] No valid GPS fix rows found — cannot build time anchors")
        return []

    anchors = []
    for _, row in df_valid.iterrows():
        try:
            tus   = int(row[tus_col])
            gwk   = int(row[gwk_col])
            gms   = int(row[gms_col])
            utc_dt = gps_week_ms_to_utc(gwk, gms)
            anchors.append((tus, utc_dt))
        except Exception:
            continue

    log(f"    GPS time anchors: {len(anchors)} valid fix rows")
    if anchors:
        t0, d0 = anchors[0]
        log(f"    First anchor: TimeUS={t0}  →  {d0.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]} UTC")
    return anchors


def timeus_to_datetime_str(target_tus, anchors, suffix=""):
    """
    Convert a log TimeUS value to a wall-clock datetime string.

    Finds the nearest GPS anchor (by TimeUS distance) and offsets from it:
        wall_clock = anchor_utc + timedelta(microseconds=(target_tus - anchor_tus))

    Returns 'YYYY-MM-DD HH:MM:SS.mmm' or falls back to raw TimeUS string.
    """
    if not anchors:
        return str(int(target_tus)) + suffix

    target = int(target_tus)

    # Find the anchor whose TimeUS is closest to target_tus
    best_anchor_tus, best_anchor_dt = min(anchors, key=lambda a: abs(a[0] - target))

    offset_us = target - best_anchor_tus
    wall_clock = best_anchor_dt + timedelta(microseconds=offset_us)

    # Shift UTC → IST (UTC +5:30)
    ist_clock = wall_clock + IST_OFFSET

    # Format: YYYY-MM-DD HH:MM:SS.mmm  (milliseconds, 3 digits)
    ms = ist_clock.microsecond // 1000
    dt_str = ist_clock.strftime("%Y-%m-%d %H:%M:%S") + f".{ms:03d}"
    return dt_str + suffix


MODE_MAP = {
    0:"Stabilize",1:"Acro",2:"AltHold",3:"Auto",4:"Guided",5:"Loiter",
    6:"RTL",7:"Circle",9:"Land",11:"Drift",13:"Sport",14:"Flip",
    15:"AutoTune",16:"PosHold",17:"Brake",18:"Throw",19:"Avoid_ADSB",
    20:"Guided_NoGPS",21:"Smart_RTL",22:"FlowHold",23:"Follow",24:"ZigZag",
}


# ═══════════════════════════════════════════════════════════════════════════
#  TAKEOFF / LANDING DETECTION  —  CTUN ThO row-by-row scan
# ═══════════════════════════════════════════════════════════════════════════

def detect_takeoff_landing(csv_dir, gps_anchors):
    """
    Scan CTUN.ThO row by row.
    TAKEOFF  = first row where ThO > 0
    LANDING  = last  row where ThO > 0, before ThO drops to 0

    Wall-clock datetime for each event is computed from GPS anchors:
        timeus_to_datetime_str(event_TimeUS, gps_anchors)

    This is the correct, accurate method. No synthetic timestamps used.
    """
    result = {
        "takeoffs":    [],
        "landings":    [],
        "first_tus":   None,
        "last_tus":    None,
        "log_end_tus": None,
        "has_partial": False,
    }

    df = load(csv_dir, "CTUN")
    if df is None:
        print("    [INFO] CTUN.csv not found"); return result

    tho_col = find_col(df, "ThO","Thr","Throttle","ThrottleOut","ThrOut","ThO_pct","tho")
    tus_col = find_col(df, "TimeUS","timeus","Time","timestamp")
    if tho_col is None or tus_col is None:
        print("    [WARN] ThO or TimeUS column missing"); return result

    df[tus_col] = pd.to_numeric(df[tus_col], errors="coerce")
    df[tho_col] = pd.to_numeric(df[tho_col], errors="coerce")
    df_clean = (df.dropna(subset=[tus_col, tho_col])
                  .sort_values(tus_col)
                  .reset_index(drop=True))

    tus_arr = df_clean[tus_col].values
    tho_arr = df_clean[tho_col].values

    if len(tus_arr) < 2:
        print("    [WARN] Not enough CTUN rows"); return result

    result["log_end_tus"] = int(tus_arr[-1])

    state            = "ON_GROUND"
    last_nonzero_tus = None

    for i in range(len(tus_arr)):
        tus = int(tus_arr[i])
        tho = float(tho_arr[i])

        if state == "ON_GROUND":
            if tho > THO_FLYING_THRESHOLD:
                takeoff_tus = tus
                last_nonzero_tus = tus
                state = "IN_FLIGHT"
                # ── Use GPS anchor to get real wall-clock time ──
                ts = timeus_to_datetime_str(takeoff_tus, gps_anchors)
                result["takeoffs"].append((takeoff_tus, ts))
                print(f"    → Takeoff #{len(result['takeoffs'])}: {ts}  "
                      f"(TimeUS={takeoff_tus}  ThO={tho:.4f})")
        else:
            if tho > THO_FLYING_THRESHOLD:
                last_nonzero_tus = tus
            else:
                landing_tus = last_nonzero_tus
                state = "ON_GROUND"
                # ── Use GPS anchor to get real wall-clock time ──
                ts = timeus_to_datetime_str(landing_tus, gps_anchors)
                result["landings"].append((landing_tus, ts))
                print(f"    → Landing #{len(result['landings'])}: {ts}  "
                      f"(TimeUS={landing_tus})")

    if state == "IN_FLIGHT" and last_nonzero_tus is not None:
        ts = timeus_to_datetime_str(last_nonzero_tus, gps_anchors, suffix=" (log end)")
        result["landings"].append((last_nonzero_tus, ts))
        result["has_partial"] = True
        print(f"    → Landing #{len(result['landings'])}: {ts}  ⚠ PARTIAL")

    if result["takeoffs"]: result["first_tus"] = result["takeoffs"][0][0]
    if result["landings"]: result["last_tus"]  = result["landings"][-1][0]
    print(f"    ✓ {len(result['takeoffs'])} takeoff(s), "
          f"{len(result['landings'])} landing(s)")
    return result


# ═══════════════════════════════════════════════════════════════════════════
#  ENDURANCE
# ═══════════════════════════════════════════════════════════════════════════

def compute_endurance(takeoffs, landings, log_end_tus=None):
    result = {
        "total_sec": 0.0, "total_str": "N/A",
        "segments": [], "has_partial": False,
        "partial_note": "", "is_valid": False,
    }
    if not takeoffs:
        result["partial_note"] = "No takeoff events detected"; return result

    total_sec = 0.0; n_complete = 0; n_partial = 0

    for idx, (to_tus, to_str) in enumerate(takeoffs):
        seg = {"n": idx+1, "takeoff_tus": to_tus, "takeoff_str": to_str,
               "landing_tus": None, "landing_str": "—",
               "seg_sec": 0.0, "seg_str": "—", "partial": False}

        if idx < len(landings):
            la_tus, la_str = landings[idx]
            seg_sec    = max(0.0, (la_tus - to_tus) / 1_000_000)
            is_partial = "(log end)" in la_str
            seg.update({"landing_tus": la_tus, "landing_str": la_str,
                        "seg_sec": seg_sec, "partial": is_partial})
            if is_partial:
                seg["seg_str"] = f"⚠ {sec_to_dur(seg_sec)} PARTIAL"
                n_partial += 1; result["has_partial"] = True
            else:
                seg["seg_str"] = sec_to_dur(seg_sec); n_complete += 1
        else:
            seg["partial"] = True; n_partial += 1; result["has_partial"] = True
            if log_end_tus:
                seg_sec = max(0.0, (log_end_tus - to_tus) / 1_000_000)
                seg.update({"landing_tus": log_end_tus,
                            "landing_str": str(log_end_tus) + " (log end)",
                            "seg_sec": seg_sec,
                            "seg_str": f"⚠ {sec_to_dur(seg_sec)} PARTIAL"})
            else:
                seg["seg_str"] = "⚠ PARTIAL"

        total_sec += seg["seg_sec"]; result["segments"].append(seg)
        flag = "  ⚠ PARTIAL" if seg["partial"] else ""
        print(f"      Seg {idx+1}: {seg['takeoff_str']} → "
              f"{seg['landing_str']}  =  {seg['seg_str']}{flag}")

    result["total_sec"] = total_sec
    result["total_str"] = sec_to_dur(total_sec)
    result["is_valid"]  = True
    result["partial_note"] = (
        f"⚠ {n_partial} partial segment(s) — log ended mid-flight. Included."
        if result["has_partial"] else f"✅ {n_complete} complete segment(s)"
    )
    print(f"    ✓ Endurance: {result['total_str']}  "
          f"({n_complete} complete, {n_partial} partial)")
    return result


# ═══════════════════════════════════════════════════════════════════════════
#  VIBRATION — 500 ms sampled, on-air only, per axis independently
# ═══════════════════════════════════════════════════════════════════════════

def compute_vibe(csv_dir, takeoffs, landings):
    """
    Compute on-air vibration averages using 500 ms interval sampling.
    (Unchanged from v3.4 — vibration does not need GPS time.)
    """
    EMPTY = {"mean": None, "max": None, "band": "N/A",
             "n_samples": 0, "n_segs": 0}

    out = {
        "X": dict(EMPTY), "Y": dict(EMPTY), "Z": dict(EMPTY),
        "vibe_window":   "no VIBE data",
        "vibe_above_20": False,
    }

    df_vibe = load(csv_dir, "VIBE")
    if df_vibe is None:
        out["vibe_window"] = "VIBE.csv not found"; return out

    if not takeoffs:
        out["vibe_window"] = "no takeoff events"; return out

    vt_col = find_col(df_vibe, "TimeUS", "timeus", "Time_us")
    if vt_col is None:
        out["vibe_window"] = "no TimeUS in VIBE.csv"; return out

    df_vibe = df_vibe.copy()
    df_vibe[vt_col] = pd.to_numeric(df_vibe[vt_col], errors="coerce")
    df_vibe = (df_vibe.dropna(subset=[vt_col])
                      .sort_values(vt_col)
                      .reset_index(drop=True))
    vibe_tus = df_vibe[vt_col].values.astype("int64")
    n_vibe   = len(vibe_tus)

    ax_vals = {}
    for ax in ("X", "Y", "Z"):
        col = find_col(df_vibe, f"Vibe{ax}", f"vibe{ax}", f"vibe_{ax.lower()}")
        if col is not None:
            raw = pd.to_numeric(df_vibe[col], errors="coerce")
            ax_vals[ax] = raw.abs().values
        else:
            ax_vals[ax] = None
            print(f"    [WARN] Vibe{ax} column not found in VIBE.csv")

    segments = []
    for idx, (to_tus, _) in enumerate(takeoffs):
        if idx < len(landings):
            la_tus, _ = landings[idx]
            segments.append((idx + 1, int(to_tus), int(la_tus)))

    if not segments:
        out["vibe_window"] = "no matched takeoff/landing pairs"; return out

    pool = {"X": [], "Y": [], "Z": []}
    seg_summaries = []

    for seg_n, to_tus, la_tus in segments:
        targets = range(to_tus, la_tus + 1, VIBE_SAMPLE_INTERVAL)
        selected_in_seg = set()
        n_selected = 0

        for target_tus in targets:
            ins = int(pd.Series(vibe_tus).searchsorted(target_tus))
            best_idx  = None
            best_diff = None

            for candidate in (ins, ins - 1):
                if candidate < 0 or candidate >= n_vibe:
                    continue
                t = int(vibe_tus[candidate])
                if not (to_tus <= t <= la_tus):
                    continue
                diff = abs(t - target_tus)
                if best_diff is None or diff < best_diff:
                    best_diff = diff
                    best_idx  = candidate

            if best_idx is None or best_idx in selected_in_seg:
                continue

            selected_in_seg.add(best_idx)
            n_selected += 1

            for ax in ("X", "Y", "Z"):
                if ax_vals[ax] is None:
                    continue
                v = float(ax_vals[ax][best_idx])
                if not pd.isna(v):
                    pool[ax].append(v)

        seg_summaries.append(f"seg{seg_n}({n_selected} samples)")
        print(f"    On-air seg {seg_n}: [{to_tus} … {la_tus}]  "
              f"targets={len(list(range(to_tus, la_tus+1, VIBE_SAMPLE_INTERVAL)))}  "
              f"selected={n_selected}")

    out["vibe_window"] = (
        f"{len(segments)} segment(s), 500ms nearest-value sampling: "
        + ", ".join(seg_summaries)
    )

    for ax in ("X", "Y", "Z"):
        p = pool[ax]
        if not p:
            out[ax] = dict(EMPTY)
            out[ax]["n_segs"] = len(segments)
            print(f"    Vibe{ax}: 0 valid samples")
            continue

        avg_val = round(sum(p) / len(p), 4)
        max_val = round(max(p), 4)

        band = ("CRITICAL" if max_val > VIBE_WARNING else
                "WARNING"  if max_val > VIBE_CAUTION else
                "CAUTION"  if max_val > VIBE_OK      else
                "OK"       if max_val > VIBE_GOOD     else "GOOD")

        out[ax] = {
            "mean":      avg_val,
            "max":       max_val,
            "band":      band,
            "n_samples": len(p),
            "n_segs":    len(segments),
        }

        if max_val >= VIBE_CAUTION:
            out["vibe_above_20"] = True

        print(f"    Vibe{ax}: n={len(p)}  avg={avg_val}  max={max_val}  [{band}]")

    return out


# ═══════════════════════════════════════════════════════════════════════════
#  CORE EXTRACTOR
# ═══════════════════════════════════════════════════════════════════════════

def extract_flight_data(csv_dir, bin_name, flight_num):
    csv_dir = Path(csv_dir)
    d = {"flight_num": flight_num, "bin_name": bin_name,
         "date": datetime.now().strftime("%d-%m-%Y")}

    # Flight mode
    df_mode  = load(csv_dir, "MODE"); mode_str = "N/A"
    if df_mode is not None:
        mc = find_col(df_mode, "Mode", "mode", "ModeNum")
        if mc is not None:
            try:
                seen = []
                for v in df_mode[mc].dropna():
                    nm = MODE_MAP.get(int(v), f"Mode{int(v)}")
                    if nm not in seen: seen.append(nm)
                mode_str = " / ".join(seen) if seen else "N/A"
            except Exception: pass
    d["flight_mode"] = mode_str

    # ── Build GPS time anchors FIRST ───────────────────────────────────────
    # This is the key fix: real wall-clock times come from GPS, not _timestamp
    print("    Building GPS time anchors…")
    gps_anchors = build_gps_time_anchors(csv_dir)
    if not gps_anchors:
        print("    [WARN] No GPS anchors — takeoff/landing times will show raw TimeUS")
    d["_gps_anchors"] = gps_anchors

    # Takeoff / Landing — datetime derived from GPS anchors
    print("    Detecting takeoff/landing from CTUN.ThO…")
    tl = detect_takeoff_landing(csv_dir, gps_anchors)
    d["takeoffs"]      = tl["takeoffs"]
    d["landings"]      = tl["landings"]
    d["takeoff_count"] = len(tl["takeoffs"])
    d["landing_count"] = len(tl["landings"])
    d["log_end_tus"]   = tl["log_end_tus"]

    first_tus = tl["first_tus"]
    last_tus  = tl["last_tus"]

    if first_tus is not None and last_tus is not None:
        dur = (last_tus - first_tus) / 1_000_000
        d["takeoff_time"] = tl["takeoffs"][0][1]  if tl["takeoffs"] else "N/A"
        d["land_time"]    = tl["landings"][-1][1] if tl["landings"] else "N/A"
        d["flight_dur"]   = sec_to_dur(dur)
        d["flight_min"]   = int(dur // 60)
        d["flight_sec"]   = int(dur % 60)
    else:
        d["takeoff_time"] = d["land_time"] = d["flight_dur"] = "N/A"
        d["flight_min"]   = d["flight_sec"] = None

    d["_first_tus"] = first_tus
    d["_last_tus"]  = last_tus

    # Endurance
    print("    Computing endurance…")
    end = compute_endurance(d["takeoffs"], d["landings"], d["log_end_tus"])
    d["endurance"]         = end
    d["endurance_str"]     = end["total_str"]
    d["endurance_sec"]     = end["total_sec"]
    d["endurance_partial"] = end["has_partial"]
    d["endurance_note"]    = end["partial_note"]

    # Battery
    df_batt = load(csv_dir, "BATT", "BAT", "CURR")
    batt_volt_max = batt_volt_min = batt_volt_takeoff = batt_volt_land = None
    batt_curr_max = batt_curr_mean = None
    volt_drop_events = curr_spike_events = 0
    batt_volt_status = "OK"
    if df_batt is not None:
        vc = find_col(df_batt, "Volt","volt","Voltage","voltage","Vcc","VBat")
        cc = find_col(df_batt, "Curr","curr","Current","current","IBat")
        if vc is not None:
            s = df_batt[vc].dropna()
            batt_volt_max     = round(float(s.max()), 3)
            batt_volt_min     = round(float(s.min()), 3)
            n = min(20, len(s))
            batt_volt_takeoff = round(float(s.iloc[:n].mean()), 3)
            batt_volt_land    = round(float(s.iloc[-n:].mean()), 3)
            volt_drop_events  = int((s.diff() < -VOLT_DROP_THR).sum())
            if batt_volt_min < BATT_VOLT_CRIT: batt_volt_status = "CRITICAL"
            elif batt_volt_min < BATT_VOLT_WARN: batt_volt_status = "WARNING"
        if cc is not None:
            s = df_batt[cc].dropna()
            batt_curr_max     = round(float(s.max()), 2)
            batt_curr_mean    = round(float(s.mean()), 2)
            curr_spike_events = int((s > BATT_CURR_HIGH).sum())
    d.update({
        "batt_volt_max": batt_volt_max, "batt_volt_min": batt_volt_min,
        "batt_volt_takeoff": batt_volt_takeoff, "batt_volt_land": batt_volt_land,
        "batt_curr_max": batt_curr_max, "batt_curr_mean": batt_curr_mean,
        "volt_drop_events": volt_drop_events, "curr_spike_events": curr_spike_events,
        "batt_volt_status": batt_volt_status,
    })

    # Vibration — 500ms sampled, on-air only, per axis
    print("    Computing vibration (500ms sampling, on-air windows)…")
    vr = compute_vibe(csv_dir, d["takeoffs"], d["landings"])
    d["vibe"]          = {ax: vr[ax] for ax in ("X", "Y", "Z")}
    d["vibe_above_20"] = vr["vibe_above_20"]
    d["vibe_window"]   = vr["vibe_window"]
    d["vibe_note"]     = (
        "⚠ Vibe >=20 m/s2 on "
        + "/".join(ax for ax in ("X","Y","Z")
                   if (vr[ax].get("max") or 0) >= VIBE_CAUTION)
        if vr["vibe_above_20"] else "All axes OK"
    )

    # Altitude
    df_baro = load(csv_dir, "BARO"); baro_max = baro_min = None
    if df_baro is not None:
        ac = find_col(df_baro, "Alt", "alt", "altitude")
        if ac is not None:
            s = df_baro[ac].dropna()
            baro_max = round(float(s.max()), 1)
            baro_min = round(float(s.min()), 1)
    d["baro_alt_max"] = baro_max
    d["baro_alt_min"] = baro_min
    return d


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════

def _cell(ws, row, col, value=None, fill_c=None, font=None,
          align=AL_C, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(fill_c or C_WHITE); c.font = font or FNT_BODY
    c.alignment = align; c.border = thin()
    if num_fmt: c.number_format = num_fmt
    return c

def _vibe_fill_font(band):
    return {"GOOD":(C_GOOD,FNT_GREEN),"OK":(C_OK,FNT_GREEN),
            "CAUTION":(C_CAUTION,FNT_YELLOW),"WARNING":(C_WARNING,FNT_ORANGE),
            "CRITICAL":(C_CRITICAL,FNT_WHITE),"N/A":(C_WHITE,FNT_BODY)
            }.get(band, (C_WHITE, FNT_BODY))


def build_excel(flights, output_path):
    wb = Workbook(); ws = wb.active
    ws.title = "Flight Log"; ws.freeze_panes = "A5"

    max_to = max(max((fd["takeoff_count"] for fd in flights), default=1), 1)
    max_la = max(max((fd["landing_count"] for fd in flights), default=1), 1)

    C_INFO_S=1; C_INFO_E=4; C_TIME_S=5; C_TIME_E=8; C_ENDUR=8
    C_TO_S=9;   C_TO_E=C_TO_S+max_to-1
    C_LA_S=C_TO_E+1; C_LA_E=C_LA_S+max_la-1
    C_BATT_S=C_LA_E+1; C_BATT_E=C_BATT_S+6
    C_ALT_S=C_BATT_E+1; C_ALT_E=C_ALT_S+1
    C_VX_S=C_ALT_E+1;  C_VX_E=C_VX_S+2
    C_VY_S=C_VX_E+1;   C_VY_E=C_VY_S+2
    C_VZ_S=C_VY_E+1;   C_VZ_E=C_VZ_S+2
    C_REM_S=C_VZ_E+1;  C_REM_E=C_REM_S+1
    TOTAL=C_REM_E

    def mhdr(row, cs, ce, text, colour, font=FNT_HDR):
        if ce > cs: ws.merge_cells(start_row=row, start_column=cs,
                                   end_row=row, end_column=ce)
        c = ws.cell(row=row, column=cs, value=text)
        c.fill=fill(colour); c.font=font; c.alignment=AL_C; c.border=thick()
        for ci in range(cs+1, ce+1): ws.cell(row=row, column=ci).border=thin()
        return c

    # ROW 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL)
    tc = ws.cell(row=1, column=1,
                 value=f"Flight Log Report  v3.6  —  Generated: "
                       f"{datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
                       f"  |  Times in IST (UTC +5:30) via GPS GWk+GMS anchors")
    tc.fill=fill(C_HDR_DARK); tc.font=Font(name="Arial",bold=True,color=C_WHITE,size=13)
    tc.alignment=AL_C; tc.border=thick(); ws.row_dimensions[1].height=30

    # ROW 2: Group headers
    mhdr(2,C_INFO_S,C_INFO_E,"Flight Info",C_HDR_DARK)
    mhdr(2,C_TIME_S,C_TIME_E,"Basic Timing & Endurance",C_HDR_DARK)
    mhdr(2,C_TO_S,C_TO_E,
         f"Takeoff Events  (YYYY-MM-DD HH:MM:SS.mmm IST — GPS-anchored, first ThO>0)  [{max_to}]",
         C_TAKEOFF_H)
    mhdr(2,C_LA_S,C_LA_E,
         f"Landing Events  (YYYY-MM-DD HH:MM:SS.mmm IST — GPS-anchored, last ThO>0)  [{max_la}]",
         C_LANDING_H)
    mhdr(2,C_BATT_S,C_BATT_E,"Battery","154360")
    mhdr(2,C_ALT_S,C_ALT_E,"Altitude (m)","186A0A")
    mhdr(2,C_VX_S,C_VZ_E,
         "Vibration (m/s2)  ON-AIR ONLY — 500ms nearest-value sampling — "
         "Avg & Max per axis independently  "
         "<=10 GOOD | <=15 OK | <=20 CAUTION | <=30 WARNING | >30 CRITICAL",
         C_HDR_GREEN)
    mhdr(2,C_REM_S,C_REM_E,"Remarks",C_HDR_DARK)
    ws.row_dimensions[2].height=22

    # ROW 3: Sub-group headers
    mhdr(3,C_INFO_S,C_INFO_E,"",C_HDR_DARK,FNT_SUBHDR)
    mhdr(3,C_TIME_S,C_TIME_S+2,"Timing (first TO → last LD)",C_HDR_DARK,FNT_SUBHDR)
    mhdr(3,C_ENDUR,C_ENDUR,"Endurance = Σ(landing_N−takeoff_N)",C_ENDUR_H,FNT_SUBHDR)
    mhdr(3,C_TO_S,C_TO_E,
         "GPS-anchored IST (UTC+5:30)  (GWk+GMS → UTC, +5h30m applied)",
         C_TAKEOFF_H,FNT_SUBHDR)
    mhdr(3,C_LA_S,C_LA_E,
         "GPS-anchored IST (UTC+5:30)  (GWk+GMS → UTC, +5h30m applied)",
         C_LANDING_H,FNT_SUBHDR)
    mhdr(3,C_BATT_S,C_BATT_S+1,"Voltage (V)","1A6B7A",FNT_SUBHDR)
    mhdr(3,C_BATT_S+2,C_BATT_S+3,"Volt @ Event","1A6B7A",FNT_SUBHDR)
    mhdr(3,C_BATT_S+4,C_BATT_S+4,"Drop","1A6B7A",FNT_SUBHDR)
    mhdr(3,C_BATT_S+5,C_BATT_E,"Current (A)","1A6B7A",FNT_SUBHDR)
    mhdr(3,C_ALT_S,C_ALT_E,"","186A0A",FNT_SUBHDR)
    mhdr(3,C_VX_S,C_VX_E,"VibeX  on-air 500ms avg",C_HDR_GREEN,FNT_SUBHDR)
    mhdr(3,C_VY_S,C_VY_E,"VibeY  on-air 500ms avg",C_HDR_GREEN,FNT_SUBHDR)
    mhdr(3,C_VZ_S,C_VZ_E,"VibeZ  on-air 500ms avg",C_HDR_GREEN,FNT_SUBHDR)
    mhdr(3,C_REM_S,C_REM_E,"",C_HDR_DARK,FNT_SUBHDR)
    ws.row_dimensions[3].height=18

    # ROW 4: Column headers
    def chdr(col, label, width, colour=C_HDR_MID, fnt=FNT_SUBHDR):
        c = ws.cell(row=4, column=col, value=label)
        c.fill=fill(colour); c.font=fnt; c.alignment=AL_C; c.border=thin()
        ws.column_dimensions[get_column_letter(col)].width=width

    chdr(1,"Flight #",8); chdr(2,"BIN File Name",22); chdr(3,"Date",13)
    chdr(4,"Flight Mode",20)
    chdr(5,"First Takeoff\nYYYY-MM-DD HH:MM:SS.mmm (IST)", 28)
    chdr(6,"Last Landing\nYYYY-MM-DD HH:MM:SS.mmm (IST)",  28)
    chdr(7,"Duration\n(1st→last)", 13)
    chdr(C_ENDUR,"Endurance\n(Total Airtime)",16,colour=C_ENDUR_H,fnt=FNT_SUBHDR)

    for i in range(max_to):
        chdr(C_TO_S+i,
             f"Takeoff_{i+1}\nYYYY-MM-DD HH:MM:SS.mmm IST\n(GPS-anchored, first ThO>0)", 28,
             colour=C_TAKEOFF_H, fnt=FNT_SUBHDR)
    for i in range(max_la):
        chdr(C_LA_S+i,
             f"Landing_{i+1}\nYYYY-MM-DD HH:MM:SS.mmm IST\n(GPS-anchored, last ThO>0)", 28,
             colour=C_LANDING_H, fnt=FNT_SUBHDR)

    chdr(C_BATT_S,"Max (V)",10); chdr(C_BATT_S+1,"Min (V)",10)
    chdr(C_BATT_S+2,"Takeoff\nVolt",11); chdr(C_BATT_S+3,"Land\nVolt",11)
    chdr(C_BATT_S+4,"Drop\nEvents",10); chdr(C_BATT_S+5,"Curr\nMax (A)",10)
    chdr(C_BATT_S+6,"Curr\nMean (A)",10)
    chdr(C_ALT_S,"Alt Min (m)",10); chdr(C_ALT_S+1,"Alt Max (m)",10)

    for base, ax in [(C_VX_S,"X"),(C_VY_S,"Y"),(C_VZ_S,"Z")]:
        chdr(base,   f"Vibe{ax}\nOn-Air Avg\n(m/s2)", 14,
             colour=C_HDR_GREEN, fnt=FNT_SUBHDR)
        chdr(base+1, f"Vibe{ax}\nMax\n(m/s2)", 14,
             colour=C_HDR_GREEN, fnt=FNT_SUBHDR)
        chdr(base+2, f"Vibe{ax}\nStatus", 11,
             colour=C_HDR_GREEN, fnt=FNT_SUBHDR)

    chdr(C_REM_S,"Vibe Note",32); chdr(C_REM_S+1,"Endurance Note",36)
    ws.row_dimensions[4].height=44

    # DATA ROWS
    for ri, fd in enumerate(flights):
        row = ri + 5
        bg  = C_ALT if ri % 2 == 0 else C_WHITE

        def dc(col, value, fc=None, font=None, align=AL_C, fmt=None):
            _cell(ws, row, col, value, fc or bg, font or FNT_BODY, align, fmt)

        dc(1,fd["flight_num"],font=FNT_BOLD); dc(2,fd["bin_name"],align=AL_L)
        dc(3,fd["date"]); dc(4,fd["flight_mode"],align=AL_L)
        dc(5,fd["takeoff_time"], align=AL_L)
        dc(6,fd["land_time"],    align=AL_L)
        dc(7,fd["flight_dur"])

        is_partial = fd["endurance_partial"]
        end_val = fd["endurance_str"] + (" (PARTIAL)" if is_partial else "")
        dc(C_ENDUR,end_val,
           fc=C_PARTIAL if is_partial else C_ENDUR,
           font=FNT_ORANGE if is_partial else FNT_GREEN)

        for i in range(max_to):
            col = C_TO_S+i
            if i < len(fd["takeoffs"]):
                dc(col, fd["takeoffs"][i][1], fc=C_TAKEOFF, font=FNT_BOLD, align=AL_L)
            else:
                dc(col, "—", fc=C_TAKEOFF)

        for i in range(max_la):
            col = C_LA_S+i
            if i < len(fd["landings"]):
                la_str = fd["landings"][i][1]
                is_pl  = "(log end)" in la_str
                dc(col, la_str,
                   fc=C_PARTIAL if is_pl else C_LANDING,
                   font=FNT_ORANGE if is_pl else FNT_BOLD,
                   align=AL_L)
            else:
                dc(col, "—", fc=C_LANDING)

        vs=fd["batt_volt_status"]
        v_fc=(C_VOLT_CRIT if vs=="CRITICAL" else C_VOLT_WARN if vs=="WARNING" else bg)
        v_fnt=(FNT_RED if vs=="CRITICAL" else FNT_ORANGE if vs=="WARNING" else FNT_BODY)
        dc(C_BATT_S,fd["batt_volt_max"],fc=v_fc,font=v_fnt,fmt="0.000")
        dc(C_BATT_S+1,fd["batt_volt_min"],fc=v_fc,font=v_fnt,fmt="0.000")
        dc(C_BATT_S+2,fd["batt_volt_takeoff"],fmt="0.000")
        dc(C_BATT_S+3,fd["batt_volt_land"],fmt="0.000")
        vde=fd["volt_drop_events"]
        dc(C_BATT_S+4,vde,
           fc=C_VOLT_CRIT if vde>10 else C_VOLT_WARN if vde>0 else bg,
           font=FNT_RED if vde>10 else FNT_ORANGE if vde>0 else FNT_BODY)
        cse=fd["curr_spike_events"]
        dc(C_BATT_S+5,fd["batt_curr_max"],
           fc=C_CURR_HIGH if cse>0 else bg,
           font=FNT_ORANGE if cse>0 else FNT_BODY,fmt="0.00")
        dc(C_BATT_S+6,fd["batt_curr_mean"],fmt="0.00")
        dc(C_ALT_S,fd["baro_alt_min"],fmt="0.0"); dc(C_ALT_S+1,fd["baro_alt_max"],fmt="0.0")

        for base, ax in [(C_VX_S,"X"),(C_VY_S,"Y"),(C_VZ_S,"Z")]:
            vd   = fd["vibe"].get(ax, {})
            avg  = vd.get("mean")
            mx   = vd.get("max")
            band = vd.get("band", "N/A")
            vfc, vfnt = _vibe_fill_font(band)
            if avg is None or mx is None:
                dc(base,"N/A"); dc(base+1,"N/A"); dc(base+2,"N/A")
            else:
                dc(base, avg,
                   fc=vfc if mx>=VIBE_CAUTION else bg,
                   font=vfnt if mx>=VIBE_CAUTION else FNT_BOLD, fmt="0.0000")
                dc(base+1, mx,
                   fc=vfc if mx>=VIBE_CAUTION else bg,
                   font=vfnt if mx>=VIBE_CAUTION else FNT_BODY, fmt="0.0000")
                dc(base+2, band, fc=vfc, font=vfnt)

        note_fc  = C_CAUTION if fd["vibe_above_20"] else bg
        note_fnt = FNT_YELLOW if fd["vibe_above_20"] else FNT_GREEN
        dc(C_REM_S,fd["vibe_note"],fc=note_fc,font=note_fnt,align=AL_L)
        dc(C_REM_S+1,fd["endurance_note"],align=AL_L,
           fc=C_PARTIAL if fd["endurance_partial"] else bg,
           font=FNT_ORANGE if fd["endurance_partial"] else FNT_BODY)
        ws.row_dimensions[row].height=22

    # LEGEND sheet
    wl = wb.create_sheet("Legend")
    wl.column_dimensions["A"].width=30
    wl.column_dimensions["B"].width=32
    wl.column_dimensions["C"].width=62

    legend_rows = [
        ("DATETIME METHOD (v3.6)","",""),
        ("Source","GPS message fields GWk + GMS","GPS week + milliseconds into week"),
        ("Epoch","1980-01-06 00:00:00 UTC","Standard GPS epoch"),
        ("Leap seconds","18 seconds subtracted","GPS − UTC offset as of 2017"),
        ("Timezone","UTC + 5h 30m = IST","Applied at final formatting step"),
        ("Anchor selection","Nearest GPS fix by TimeUS","Any row with Status >= 3 and GWk > 0"),
        ("Offset formula","anchor_utc + timedelta(µs = ThO_TimeUS − anchor_TimeUS) + 5h30m",
         "UTC computed first, then IST offset added"),
        ("Output format","YYYY-MM-DD HH:MM:SS.mmm IST","Millisecond precision, India Standard Time"),
        ("","",""),
        ("VIBRATION LOGIC (v3.5)","",""),
        ("Rule","Value","Description"),
        ("On-air only","Takeoff_N → Landing_N",
         "Only VIBE rows inside ThO>0 windows are used. Ground data is ignored."),
        ("500ms sampling","Every 500,000 µs",
         "Target timestamps spaced 500ms apart starting at takeoff_tus. "
         "Nearest VIBE row selected per target."),
        ("Per-axis average","sum(|Vibe|) / count",
         "Axes X, Y, Z computed independently — never averaged together."),
        ("","",""),
        ("TAKEOFF / LANDING LOGIC","",""),
        ("Takeoff_N","First CTUN row where ThO > 0","TimeUS used as GPS offset anchor."),
        ("Landing_N","Last CTUN row where ThO > 0, before drop","TimeUS used as GPS offset anchor."),
        ("","",""),
        ("ENDURANCE CALCULATION","",""),
        ("Segment","landing_N.tus − takeoff_N.tus","µs → seconds per pair."),
        ("Total","Σ all segment seconds","Complete + partial summed."),
        ("","",""),
        ("VIBRATION COLOUR SCALE","",""),
        ("Band","Max (m/s2)","Meaning"),
        ("GOOD","<= 10","Healthy"),("OK","10–15","Acceptable"),
        ("CAUTION","15–20","Inspect motors/props"),
        ("WARNING","20–30","Fix required"),("CRITICAL","> 30","Crash risk"),
        ("","",""),
        ("BATTERY COLOUR SCALE","",""),
        ("Volt OK",f"> {BATT_VOLT_WARN} V","Normal"),
        ("Volt WARN",f"< {BATT_VOLT_WARN} V","Low"),
        ("Volt CRIT",f"< {BATT_VOLT_CRIT} V","Critical"),
        ("Curr spike","Any row",f"Current > {BATT_CURR_HIGH} A"),
        ("Volt drop","Drop events",f"Voltage drop > {VOLT_DROP_THR} V/sample"),
    ]
    clr_map = {
        "GOOD":C_GOOD,"OK":C_OK,"CAUTION":C_CAUTION,"WARNING":C_WARNING,"CRITICAL":C_CRITICAL,
        "Volt OK":C_WHITE,"Volt WARN":C_VOLT_WARN,"Volt CRIT":C_VOLT_CRIT,
        "Curr spike":C_CURR_HIGH,"Volt drop":C_VOLT_WARN,
        "On-air only":C_ENDUR,"500ms sampling":C_ALT,
        "Per-axis average":C_GOOD,"Max":C_OK,
    }
    dark_hdrs = {"DATETIME METHOD (v3.6)","VIBRATION LOGIC (v3.5)",
                 "TAKEOFF / LANDING LOGIC","ENDURANCE CALCULATION",
                 "VIBRATION COLOUR SCALE","BATTERY COLOUR SCALE"}
    for ri,(a,b,c) in enumerate(legend_rows,1):
        for ci,val in enumerate([a,b,c],1):
            cell=wl.cell(row=ri,column=ci,value=val)
            cell.border=thin(); cell.alignment=AL_L
            cell.font=FNT_BOLD if ci==1 else FNT_BODY
            if a in clr_map and ci==1: cell.fill=fill(clr_map[a])
            if a in dark_hdrs: cell.fill=fill(C_HDR_DARK); cell.font=FNT_HDR
            elif a in ("Band","Rule","Source"): cell.fill=fill(C_HDR_MID); cell.font=FNT_SUBHDR
        wl.row_dimensions[ri].height=18

    wb.save(str(output_path))
    print(f"  Saved: {output_path}")
    return str(output_path)


# ═══════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

def main():
    if not (PANDAS_OK and OPENPYXL_OK):
        print("[ERROR] pip install pandas openpyxl numpy pymavlink"); sys.exit(1)

    bin_paths = []
    if len(sys.argv) > 1:
        for arg in sys.argv[1:]:
            p = Path(arg)
            if p.is_dir():
                found = list(p.glob("**/*.bin")) + list(p.glob("**/*.BIN"))
                bin_paths.extend(found)
                print(f"  Found {len(found)} .bin file(s) in {p}")
            elif p.suffix.lower() == ".bin" and p.exists():
                bin_paths.append(p)
            else:
                print(f"  [WARN] Skipping: {p}")
    elif TK_OK:
        tk_root = tk.Tk(); tk_root.withdraw()
        chosen = filedialog.askopenfilenames(
            title="Select ArduPilot .bin log file(s)",
            filetypes=[("ArduPilot Log","*.bin *.BIN"),("All","*.*")])
        tk_root.destroy(); bin_paths = [Path(p) for p in chosen]
    else:
        print("Usage: python log_report_v3_5.py <file.bin> [...]"); sys.exit(0)

    if not bin_paths:
        print("No .bin files selected."); sys.exit(0)

    print(f"\n{'='*65}")
    print(f"  Flight Log Report Generator v3.6")
    print(f"  Datetime : GPS GWk+GMS → UTC → IST (UTC+5:30)")
    print(f"  Takeoff  : first CTUN row where ThO > 0")
    print(f"  Landing  : last  CTUN row where ThO > 0, before drop")
    print(f"  Vibration: 500ms nearest-value sampling, on-air windows only")
    print(f"             avg_X, avg_Y, avg_Z computed independently")
    print(f"  Endurance= Σ (landing_N − takeoff_N)")
    print(f"  Files    : {len(bin_paths)}")
    print(f"{'='*65}")

    output_root = bin_paths[0].parent / "flight_reports"
    output_root.mkdir(parents=True, exist_ok=True)
    flights = []; failed = []

    for i, bin_path in enumerate(bin_paths, 1):
        print(f"\n[{i}/{len(bin_paths)}]  {bin_path.name}")
        print("─" * 65)
        flight_out = output_root / bin_path.stem
        flight_out.mkdir(parents=True, exist_ok=True)

        print("  Step 1/2 — Converting BIN → CSV…")
        csv_dir, csv_files = bin_to_csv(bin_path, flight_out, log=print)
        if csv_dir is None or not csv_files:
            print("  [ERROR] Conversion failed — skipping")
            failed.append(bin_path.name); continue

        important = {"BARO","BATT","CTUN","GPS","VIBE","MODE"}
        found_imp = [t for t in sorted(csv_files) if t in important]
        print(f"  Key types : {', '.join(found_imp)}"
              f"  (+{len(csv_files)-len(found_imp)} others)")

        print("  Step 2/2 — Analysing…")
        fd = extract_flight_data(csv_dir, bin_path.name, i)
        flights.append(fd)

        n_anchors = len(fd.get("_gps_anchors", []))
        print(f"  GPS anchors used: {n_anchors}")
        print(f"  Mode      : {fd['flight_mode']}")
        print(f"  Duration  : {fd['flight_dur']}"
              f"  ({fd['takeoff_time']} → {fd['land_time']})")
        print(f"  Endurance : {fd['endurance_str']}"
              f"  {'[PARTIAL]' if fd['endurance_partial'] else '[complete]'}")
        print(f"  Takeoffs  : {fd['takeoff_count']}")
        for n, (tus, ts) in enumerate(fd["takeoffs"], 1):
            print(f"    Takeoff_{n}: {ts}")
        print(f"  Landings  : {fd['landing_count']}")
        for n, (tus, ts) in enumerate(fd["landings"], 1):
            flag = "  ⚠ PARTIAL" if "(log end)" in ts else ""
            print(f"    Landing_{n}: {ts}{flag}")
        for ax in ("X","Y","Z"):
            v = fd["vibe"].get(ax, {})
            print(f"  Vibe{ax}     : avg={v.get('mean')}  max={v.get('max')}"
                  f"  n={v.get('n_samples')}  [{v.get('band','N/A')}]")

    if not flights:
        print("\n[ERROR] No flights processed."); sys.exit(1)

    print(f"\n{'='*65}")
    print(f"  Building Excel for {len(flights)} flight(s)…")
    xlsx = output_root / "flight_log_report_v3.6.xlsx"
    build_excel(flights, xlsx)

    if failed: print(f"\n  [WARN] Failed: {', '.join(failed)}")
    print(f"\n  Output → {output_root}")

    if sys.platform == "win32": os.startfile(str(output_root))
    elif sys.platform == "darwin": os.system(f'open "{output_root}"')
    else: os.system(f'xdg-open "{output_root}"')


if __name__ == "__main__":
    main()